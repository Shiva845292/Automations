import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from datetime import datetime
import os



def process_pcras_excel(file_path, output_sheet_name='PCRAS Data'):
    print('PCRAS Data')
    expected_columns = [
        "Associated User",
        "Machine Name",
        "Delivery Group",
        "Session Start Time",
        "Session End Time",
        "Session Duration (hr:min)",
        "Session Auto Reconnect Count"
    ]

    # Only remove "Delivery Group"
    columns_to_remove = ["Delivery Group"]

    try:
        df = pd.read_excel(file_path, header=None)
        header_row_index = None
        for i in range(len(df)):
            row = df.iloc[i].astype(str).str.strip().tolist()
            if all(col in row for col in expected_columns):
                header_row_index = i
                break

        if header_row_index is not None:
            data = pd.read_excel(file_path, header=header_row_index)
            filtered_data = data[expected_columns]

            # Remove only the "Delivery Group" column
            filtered_data_cleaned = filtered_data.drop(columns=columns_to_remove, errors='ignore')

            # Remove duplicate Associated Users
            filtered_data_cleaned = filtered_data_cleaned.drop_duplicates(subset="Associated User")

            # Clean Machine Name if it starts with "GITDIR\\"
            def clean_machine_name(val):
                if isinstance(val, str) and val.startswith("GITDIR\\"):
                    return val[len("GITDIR\\"):]
                return val

            filtered_data_cleaned['Machine Name'] = filtered_data_cleaned['Machine Name'].apply(clean_machine_name)

            # Keep the final columns
            final_columns = [
                "Machine Name",
                "Associated User",
                "Session Start Time",
                "Session End Time",
                "Session Duration (hr:min)",
                "Session Auto Reconnect Count"
            ]
            filtered_data_cleaned = filtered_data_cleaned[final_columns]

            # Save to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                filtered_data_cleaned.to_excel(writer, sheet_name=output_sheet_name, index=False)

            # Remove first sheet if not the output sheet
            wb = load_workbook(file_path)
            first_sheet = wb.sheetnames[0]
            if first_sheet != output_sheet_name:
                wb.remove(wb[first_sheet])
                wb.save(file_path)
                log(f"Deleted the first sheet '{first_sheet}' successfully.")
            else:
                log(f"First sheet is the output sheet '{output_sheet_name}', so not deleted.")

            log(f"Processed and saved cleaned data to '{output_sheet_name}' in file '{file_path}'.")
            return True
        else:
            log("Could not find the header row with all expected columns.")
            return False

    except FileNotFoundError:
        log(f"File not found: {file_path}")
        return False
    except Exception as e:
        log(f"An unexpected error occurred: {e}")
        return False

def apply_vlookup_and_paste_values(asset_file, pcd_file, asset_sheet_name='Sheet1', pcd_sheet_name='PCD Data'):
    print('PCD Data Vlookup')
    try:
        # Load and clean Asset and PCD data
        asset_df = pd.read_excel(asset_file, sheet_name=asset_sheet_name)
        pcd_df = pd.read_excel(pcd_file, sheet_name=pcd_sheet_name)

        # Clean column names to avoid mismatch
        pcd_df.columns = pcd_df.columns.str.strip()
        asset_df.columns = asset_df.columns.str.strip()

        print("PCD Columns:", pcd_df.columns)  # Debug: check column names

        # Normalize the keys in PCD for lookup
        pcd_df['Machine Name'] = pcd_df['Machine Name'].astype(str).str.strip().str.upper()

        # Build lookup dictionary for all required columns
        pcd_lookup = {}
        for i, row in pcd_df.iterrows():
            key = row['Machine Name']
            if key not in pcd_lookup:
                pcd_lookup[key] = {
                    'Associated User': row.get('Associated User', ''),
                    'Session Start Time': row.get('Session Start Time', ''),
                    'Session End Time': row.get('Session End Time', ''),
                    'Session Duration (hr:min)': row.get('Session Duration (hr:min)', ''),
                    'Session Auto Reconnect Count': row.get('Session Auto Reconnect Count', '')
                }

        # Normalize AssetNumber for lookup
        asset_df['AssetNumber_normalized'] = asset_df['AssetNumber'].astype(str).str.strip().str.upper()

        # Map each column using the lookup dictionary
        asset_df['Associated User'] = asset_df['AssetNumber_normalized'].map(lambda x: pcd_lookup.get(x, {}).get('Associated User', ''))
        asset_df['Session Start Time'] = asset_df['AssetNumber_normalized'].map(lambda x: pcd_lookup.get(x, {}).get('Session Start Time', ''))
        asset_df['Session End Time'] = asset_df['AssetNumber_normalized'].map(lambda x: pcd_lookup.get(x, {}).get('Session End Time', ''))
        asset_df['Session Duration (hr:min)'] = asset_df['AssetNumber_normalized'].map(lambda x: pcd_lookup.get(x, {}).get('Session Duration (hr:min)', ''))
        asset_df['Session Auto Reconnect Count'] = asset_df['AssetNumber_normalized'].map(lambda x: pcd_lookup.get(x, {}).get('Session Auto Reconnect Count', ''))

        # Drop helper column
        asset_df = asset_df.drop(columns=['AssetNumber_normalized'])

        # Save the updated Asset file
        with pd.ExcelWriter(asset_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            asset_df.to_excel(writer, sheet_name=asset_sheet_name, index=False)

        log(f"✅ VLOOKUP applied and values pasted in '{asset_file}' under sheet '{asset_sheet_name}'.")
        return True

    except FileNotFoundError:
        log(f"❌ One of the files was not found:\nAsset file: {asset_file}\nPCD file: {pcd_file}")
        return False
    except KeyError as ke:
        log(f"❌ Missing expected column: {ke}")
        return False
    except Exception as e:
        log(f"❌ Unexpected error: {e}")
        return False
def copy_sheet_to_pcras(source_file, dest_file, source_sheet='Sheet1', target_sheet='PCD_Data'):
    try:
        # Read the source sheet
        source_df = pd.read_excel(source_file, sheet_name=source_sheet)

        # Filter: Keep only rows where 'Associated User' is not empty
        if 'Associated User' in source_df.columns:
            filtered_df = source_df[source_df['Associated User'].notna()]
        else:
            raise KeyError("'Associated User' column not found in the sheet.")

        # Required columns to keep (if present)
        expected_columns = [
            'AssetNumber',
            'Associated User',
            'Session Start Time',
            'Session End Time',
            'Session Duration (hr:min)',
            'Session Auto Reconnect Count'
        ]

        # Filter to only include columns that exist in the DataFrame
        columns_to_keep = [col for col in expected_columns if col in filtered_df.columns]
        filtered_df = filtered_df[columns_to_keep]

        # Save to target sheet in destination file
        with pd.ExcelWriter(dest_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            filtered_df.to_excel(writer, sheet_name=target_sheet, index=False)

        print(f"✅ Filtered rows (including additional columns) copied to '{target_sheet}' in '{dest_file}'.")
        return True

    except Exception as e:
        print(f"❌ Failed to copy filtered sheet: {e}")
        return False

def style_and_uppercase_pcras(file_path):
    try:
        wb = load_workbook(file_path)
        target_sheets = ['PCRAS Data', 'PCD_Data']

        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue

        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Apply light blue fill to header row
                for cell in sheet[1]:
                    cell.fill = header_fill

                # Convert all cell values below header to uppercase
                for row in sheet.iter_rows(min_row=2):  # Skip header
                    for cell in row:
                        if isinstance(cell.value, str):
                            cell.value = cell.value.upper()

        wb.save(file_path)
        log(f"✅ Header formatted and data converted to uppercase in sheets: {', '.join(target_sheets)}")
        return True

    except Exception as e:
        log(f"❌ Failed to apply formatting and uppercase: {e}")
        return False


# ---------- GUI ----------

def log(message):
    text_output.config(state='normal')
    text_output.insert(tk.END, message + "\n")
    text_output.see(tk.END)
    text_output.config(state='disabled')


def browse_file(entry_widget):
    filename = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")],
        title="Select Excel file"
    )
    if filename:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, filename)
from datetime import datetime
import os

def run_all():
    text_output.config(state='normal')
    text_output.delete(1.0, tk.END)
    text_output.config(state='disabled')

    pcras_file = entry_pcras.get().strip()
    pcd_file = entry_pcd.get().strip()
    asset_file = entry_asset.get().strip()

    if not all([pcras_file, pcd_file, asset_file]):
        messagebox.showerror("Error", "Please select all three Excel files.")
        return

    log("Starting processing...\n")

    if not process_pcras_excel(pcras_file, 'PCRAS Data'):
        log("Processing PCRAS file failed.\n")
        return

    if not process_pcras_excel(pcd_file, 'PCD Data'):
        log("Processing PCD file failed.\n")
        return

    if not apply_vlookup_and_paste_values(asset_file, pcd_file, 'Sheet1', 'PCD Data'):
        log("Applying VLOOKUP failed.\n")
        return

    if not copy_sheet_to_pcras(asset_file, pcras_file, 'Sheet1', 'PCD_Data'):
        log("Copying sheet failed.\n")
        return

    # 🔹 NEW STEP: Apply formatting and uppercase transformation
    if not style_and_uppercase_pcras(pcras_file):
        log("Styling and uppercasing failed.\n")
        return

    # 🔹 Save the final file with today's date in the same directory as the PCRAS file
    try:
        today_str = datetime.now().strftime("%d-%m-%Y")
        final_filename = f"DXC EMEA Users PCRAS & PCD Data {today_str}.xlsx"
        save_dir = os.path.dirname(pcras_file)  # Save in the same folder as PCRAS file
        final_filepath = os.path.join(save_dir, final_filename)

        wb = load_workbook(pcras_file)
        wb.save(final_filepath)
        log(f"\n✅ Final file saved as '{final_filepath}'.")
    except Exception as e:
        log(f"\n❌ Failed to save final file: {e}")

    log("\nAll tasks completed successfully!")

# --- Tkinter window ---
root = tk.Tk()
root.title("Excel Processing Tool")

# Window size
root.geometry("650x500")
root.resizable(False, False)

# PCRAS file
tk.Label(root, text="Select PCRAS Excel File:").pack(anchor='w', padx=10, pady=(10,0))
frame_pcras = tk.Frame(root)
frame_pcras.pack(fill='x', padx=10)
entry_pcras = tk.Entry(frame_pcras, width=70)
entry_pcras.pack(side='left', padx=(0,5), pady=5)
btn_browse_pcras = tk.Button(frame_pcras, text="Browse", command=lambda: browse_file(entry_pcras))
btn_browse_pcras.pack(side='left')

# PCD file
tk.Label(root, text="Select PCD Excel File:").pack(anchor='w', padx=10, pady=(10,0))
frame_pcd = tk.Frame(root)
frame_pcd.pack(fill='x', padx=10)
entry_pcd = tk.Entry(frame_pcd, width=70)
entry_pcd.pack(side='left', padx=(0,5), pady=5)
btn_browse_pcd = tk.Button(frame_pcd, text="Browse", command=lambda: browse_file(entry_pcd))
btn_browse_pcd.pack(side='left')

# Asset file
tk.Label(root, text="Select Asset Excel File:").pack(anchor='w', padx=10, pady=(10,0))
frame_asset = tk.Frame(root)
frame_asset.pack(fill='x', padx=10)
entry_asset = tk.Entry(frame_asset, width=70)
entry_asset.pack(side='left', padx=(0,5), pady=5)
btn_browse_asset = tk.Button(frame_asset, text="Browse", command=lambda: browse_file(entry_asset))
btn_browse_asset.pack(side='left')

# Run button
btn_run = tk.Button(root, text="Run Processing", bg="green", fg="white", font=("Arial", 12, "bold"), command=run_all)
btn_run.pack(pady=15)

# Output log
tk.Label(root, text="Status Output:").pack(anchor='w', padx=10)
text_output = scrolledtext.ScrolledText(root, height=15, width=80, state='disabled')
text_output.pack(padx=10, pady=(0,10))

root.mainloop()
